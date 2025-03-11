import streamlit as st
import pandas as pd
import os
import zipfile
from io import StringIO, BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re


# Criando a interface no Streamlit
st.set_page_config(page_title="Conversor de DTA para Excel", page_icon="📄")

st.title("Conversor de Arquivos DTA para Excel")
st.write('''Faça upload de um ou mais arquivos `.DTA` para convertê-los em `.xlsx`. Em seguida você poderá fazer o download de uma pasta
         zipada. Nomeie a pasta e extraia os arquivos para a sua máquina.
         ''')

# Função para extrair metadata (DATE e TIME) do arquivo DTA
def extract_metadata(file_content):
    date_info = None
    time_info = None

    for line in file_content.split("\n"):
        partes = line.strip().split("\t")
        if partes[0] == "DATE":
            date_info = partes[-2] if len(partes) > 1 else None
        elif partes[0] == "TIME":
            time_info = partes[-2] if len(partes) > 1 else None
        if date_info and time_info:
            break

    return date_info, time_info

# Função para extrair as colunas Zreal e Zimag
def extract_zreal_zimag(file_content):
    try:
        raw_data = file_content.replace(",", ".").split("\n")

        # Identificar a linha onde os dados começam
        data_start = None
        header_line = None
        for i, line in enumerate(raw_data):
            if "Zreal" in line and "Zimag" in line:
                header_line = i
            if re.match(r"^\s*[\d\.\-,Ee]+\s+[\d\.\-,Ee]+", line):
                data_start = i
                break

        if data_start is None:
            return None

        header = raw_data[header_line].strip().split() if header_line and header_line < data_start else None

        # Transformar os dados em um DataFrame
        raw_data_str = StringIO("\n".join(raw_data[data_start:]))
        df = pd.read_csv(raw_data_str, delimiter=r'\s+', names=header, encoding="latin1", engine="python")

        if header and "Zreal" in df.columns and "Zimag" in df.columns:
            df_corrected = df[["Zreal", "Zimag"]]
        else:
            df_corrected = df.iloc[:, [3, 4]]
            df_corrected.columns = ["Zreal", "Zimag"]

        return df_corrected
    except Exception as e:
        return None



# Upload de arquivos
uploaded_files = st.file_uploader("Selecione os arquivos .DTA", type=["DTA"], accept_multiple_files=True)

# Nome da pasta de saída
pasta_saida = st.text_input("Nome da pasta de saída", "Resultado_DTA")

if uploaded_files and pasta_saida:
    arquivos_processados = []

    # Criar pasta temporária
    pasta_zip = BytesIO()
    with zipfile.ZipFile(pasta_zip, "w", zipfile.ZIP_DEFLATED) as zipf:

        for uploaded_file in uploaded_files:
            file_content = uploaded_file.getvalue().decode("latin1")

            df_final = extract_zreal_zimag(file_content)
            date_info, time_info = extract_metadata(file_content)

            if df_final is not None:
                nome_arquivo = uploaded_file.name.replace(".DTA", ".xlsx")

                # Criar Excel em memória
                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df_final.to_excel(writer, index=False, sheet_name="Dados", startrow=4)

                # Ajustar Excel para incluir metadata
                wb = load_workbook(output)
                ws = wb.active
                ws["A1"] = "Date:"
                ws["B1"] = date_info if date_info else "Não encontrado"
                ws["A2"] = "Time:"
                ws["B2"] = time_info if time_info else "Não encontrado"
                ws["A4"] = "Zreal"
                ws["B4"] = "Zimag"

                for col in range(1, ws.max_column + 1):
                    max_length = 0
                    column_letter = get_column_letter(col)
                    for cell in ws[column_letter]:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = max_length + 2

                # Salvar as alterações no Excel
                output.seek(0)
                wb.save(output)

                # Adicionar o arquivo Excel ao ZIP
                zipf.writestr(f"{pasta_saida}/{nome_arquivo}", output.getvalue())

    # Finalizar o ZIP
    pasta_zip.seek(0)

    # Criar botão para baixar o ZIP
    st.download_button(
        label="Baixar Pasta com os Arquivos",
        data=pasta_zip,
        file_name=f"{pasta_saida}.zip",
        mime="application/zip"
    )

    st.success("Conversão concluída! Baixe a pasta compactada acima.")


st.write("""
         
        ###### Versão 7.0
            Seleciona múltiplos arquivos
            Corretamente identifica o Zreal e o Zimag de um DTA
            Coleta a informação da coluna completa
            Salva como um arquivo .xlsx
            Deixa o usuário escolher a pasta de salvamento
            Importa as informações de data e horário
            Formatação correta
""")
